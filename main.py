import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

produscts_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value #give us value of column 4 form each row form sheet
    inventory = product_list.cell(product_row, 2).value #give us value of column 2 form each row form sheet
    price = product_list.cell(product_row, 3).value #give us value of column 3 form each row form sheet
    product_num = product_list.cell(product_row, 1).value #give us value of column 1 form each row form sheet
    inventory_price = product_list.cell(product_row, 5)

#calculation for num of prod per supplier
    if supplier_name in produscts_per_supplier:
        current_num_products = produscts_per_supplier.get(supplier_name)
        produscts_per_supplier[supplier_name] = current_num_products + 1
    else:
        produscts_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #calculation prod with inv under 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    #add value for total inv proce
    inventory_price.value = inventory * price

print(produscts_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")

